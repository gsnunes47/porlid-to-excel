import pandas as pd
from openpyxl import load_workbook
import string

def numero_para_letra(numero):
    """Converte um índice numérico para uma letra de coluna no Excel, suportando letras duplas."""
    alfabeto = string.ascii_uppercase
    if numero < 26:
        return alfabeto[numero]
    else:
        return alfabeto[(numero // 26) - 1] + alfabeto[numero % 26]

file_path = 'dados_convertidos.xlsx'
output_path = 'dados_convertidos.xlsx'

wb = load_workbook(output_path)
ws = wb.active
df = pd.read_excel(file_path)

linha_a = df.columns

map = {}

for indice, titulo in enumerate(linha_a):

    if (indice + 1) == len(linha_a):

        try:

            map[titulo.split('.')[0]][1] = f'{numero_para_letra(indice)}'
            cells = map.get(linha_a[indice-1].split('.')[0])
            ws.merge_cells(f'{cells[0]}1:{cells[1]}1')
            break

        except:
            
            break

    if map.get(titulo.split('.')[0]) is not None:

        map[titulo.split('.')[0]][1] = f'{numero_para_letra(indice)}'

    else:

        try:

            cells = map.get(linha_a[indice-1].split('.')[0])
            ws.merge_cells(f'{cells[0]}1:{cells[1]}1')
        
        except:

            map[titulo] = [f'{numero_para_letra(indice)}', '']
        
        else:
            
            if map.get(titulo.split('.')[0]) is not None:

                continue

            else:

                map[titulo] = [f'{numero_para_letra(indice)}', '']

wb.save(output_path)
print(f"Arquivo Excel criado com células mescladas: {output_path}")
